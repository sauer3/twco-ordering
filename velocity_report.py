#!/usr/bin/env python3
"""
Corrected reorder report for Transparent Wine Company.

Fixes the bias in Vinosmith's stock velocity report, which divides each wine's
trailing sales by a fixed number of months regardless of how many of those
months the wine was actually in stock. Because each vintage is its own SKU and
most SKUs are out of stock for part of the year (vintage not yet arrived, or
sold through), that biases velocity low and "months of supply remaining" high --
the dangerous direction given a 12-16 week reorder lead time.

Methodology (per wine), over a user-configurable trailing window:
    active window start = max(window start, wine.created_at month)
    active window end   = last month with nonzero sales   if on_hand == 0
                          current month                    otherwise
    velocity (per month) = total sales in active window / months in active window
    months_remaining     = (on_hand + on_order) / velocity
    reorder?             = months_remaining < lead_time + safety_buffer

Sales are bucketed by delivery_at (when the wine actually left inventory).

Usage:
    python velocity_report.py                 # 12-month window, defaults
    python velocity_report.py --months 18
    python velocity_report.py --lead-time 3.5 --safety-buffer 1.0
    python velocity_report.py --output output/reorder_report.xlsx
"""

import argparse
import csv
import re
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vinosmith_client import VinosmithClient

# ---- Defaults (override via CLI) -------------------------------------------
DEFAULT_WINDOW_MONTHS = 12
DEFAULT_LEAD_TIME_MONTHS = 3.5      # midpoint of the 12-16 week lead time
DEFAULT_SAFETY_BUFFER_MONTHS = 1.0
DEFAULT_OUTPUT = "output/reorder_report.xlsx"
# Append-only log of on_hand / on_order per wine, one batch per run. Over time
# this gives us the in-stock history the API doesn't expose, so velocity can be
# divided by months the wine was actually in stock (catching "arrived but sat
# unsold" as well as "not arrived yet").
# Kept at the project root (not under output/) so it persists across machines
# and isn't lost when output/ is cleaned -- it's the accumulating in-stock memory.
INVENTORY_HISTORY_PATH = "inventory_history.csv"
# Manual list of wines we are not reordering (Vinosmith's discontinued flag is
# not exposed in its API). Plain text, one wine name per line; vintage ignored.
DISCONTINUED_PATH = "discontinued.txt"
# Velocity is measured per day-in-stock, then expressed per week and per month.
DAYS_PER_MONTH = 30.44
# Floor the in-stock span so a wine that sold on a single day doesn't divide by
# ~0 and report an absurd rate; treat anything shorter as one week of selling.
MIN_VELOCITY_DAYS = 7
# A velocity built on fewer than this many days in stock is a thin signal --
# flag it for a human glance rather than trusting the reorder math. ~6 weeks:
# less than this is a shaky basis for a 12-16 week reorder decision.
THIN_HISTORY_DAYS = 42

# Domestic producers ship far faster than the European 12-16 week lead time, so
# they get a much shorter lead time (4 weeks). Matched on producer name, lower-cased.
DOMESTIC_PRODUCERS = {"bow & arrow", "mazette", "pearce predhomme"}
DOMESTIC_LEAD_TIME_MONTHS = round(4 * 7 / DAYS_PER_MONTH, 2)  # 4 weeks ~= 0.92 mo


# ---- Small month-key helpers ------------------------------------------------
def month_key(d):
    """A comparable, sortable year-month key from a date: (year, month)."""
    return (d.year, d.month)


def parse_date(s):
    """Parse an ISO-ish date string (YYYY-MM-DD...) to a date; None on failure."""
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def months_between_inclusive(start_key, end_key):
    """Number of calendar months from start_key to end_key inclusive (>=0)."""
    if end_key < start_key:
        return 0
    return (end_key[0] - start_key[0]) * 12 + (end_key[1] - start_key[1]) + 1


def add_months(d, n):
    """Return the first day of the month that is n months after d's month."""
    total = (d.year * 12 + (d.month - 1)) + n
    return date(total // 12, total % 12 + 1, 1)


# ---- Core analysis ----------------------------------------------------------
def build_daily_sales(orders):
    """
    From supplier-order records, build:
        sales[wine_id][date] = total quantity delivered that day
    Keyed on delivery_at (falls back to order_at). Day-level detail lets us
    measure velocity over the actual days a wine was in stock, so a wine that
    arrived mid-month and sold for only a few days isn't charged a full month
    in the denominator.
    """
    sales = defaultdict(lambda: defaultdict(float))
    for row in orders:
        so = row.get("supplier_order", {})
        d = parse_date(so.get("delivery_at")) or parse_date(so.get("order_at"))
        if d is None:
            continue
        for li in row.get("line_items", []):
            wine = li.get("wine") or {}
            wid = wine.get("id")
            if wid is None:
                continue
            try:
                qty = float(li.get("quantity") or 0)
            except (TypeError, ValueError):
                qty = 0.0
            sales[str(wid)][d] += qty
    return sales


def to_float(x, default=0.0):
    try:
        return float(x)
    except (TypeError, ValueError):
        return default


def aggregate_inventory(inventory):
    """Sum on_hand / on_order per wine id across warehouses."""
    agg = defaultdict(lambda: {"on_hand": 0.0, "on_order": 0.0})
    for rec in inventory:
        wid = (rec.get("wine") or {}).get("id")
        if wid is None:
            continue
        inv = rec.get("inventory") or {}
        agg[str(wid)]["on_hand"] += to_float(inv.get("on_hand"))
        agg[str(wid)]["on_order"] += to_float(inv.get("on_order"))
    return agg


def record_snapshot(inventory, snapshot_date, path=INVENTORY_HISTORY_PATH):
    """
    Append this run's on_hand / on_order per wine to the history CSV. Re-running
    on the same snapshot_date overwrites that date's rows (idempotent per day),
    so accidental double-runs don't double-count.
    """
    agg = aggregate_inventory(inventory)
    snap = snapshot_date.isoformat()
    fields = ["snapshot_date", "wine_id", "on_hand", "on_order"]

    kept = []
    p = Path(path)
    if p.exists():
        with open(p, newline="") as f:
            for row in csv.DictReader(f):
                if row.get("snapshot_date") != snap:
                    kept.append(row)

    for wid, v in sorted(agg.items()):
        kept.append({
            "snapshot_date": snap,
            "wine_id": wid,
            "on_hand": v["on_hand"],
            "on_order": v["on_order"],
        })

    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(kept)
    return p


def load_instock_months(path=INVENTORY_HISTORY_PATH):
    """
    From the snapshot history, return {wine_id: set of (year, month)} for every
    month in which a snapshot recorded on_hand > 0 -- i.e. months we have direct
    evidence the wine was in stock.
    """
    res = defaultdict(set)
    p = Path(path)
    if not p.exists():
        return res
    with open(p, newline="") as f:
        for row in csv.DictReader(f):
            if to_float(row.get("on_hand")) > 0:
                d = parse_date(row.get("snapshot_date"))
                if d is not None:
                    res[str(row.get("wine_id"))].add(month_key(d))
    return res


# Matches a vintage token in a wine name: an apostrophe-year ('24) or a bare
# 4-digit year (2024). Used to collapse vintages of the same wine into a family.
_VINTAGE_RE = re.compile(r"'\d{2}\b|\b(?:19|20)\d{2}\b")


def strip_vintage(name):
    """A wine name with the vintage token removed and whitespace tidied."""
    stripped = _VINTAGE_RE.sub("", name or "")
    # Tidy up the gaps left behind ('Funaro -  "Passo..."  ->  'Funaro - "Passo').
    return re.sub(r"\s+", " ", stripped).strip(" -–")


def family_name(wine):
    """
    The wine's name with the vintage stripped, so different vintages of the same
    cuvee collapse to one product family. The name already starts with the
    producer (e.g. 'Funaro - \\'24 "Passo di Luna" Bianco'), and 'MAGNUM' and
    bottle sizes stay in the name, so those remain distinct families.
    """
    return strip_vintage(wine.get("name") or "") or (wine.get("name") or "")


def load_discontinued(path):
    """
    Read the manual discontinued list into a set of vintage-stripped, lower-cased
    family keys. Vinosmith's discontinued flag isn't in the API, so this file is
    how we mark wines we're not reordering. Missing file -> empty set.
    """
    keys = set()
    p = Path(path)
    if not p.exists():
        return keys
    with open(p, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            keys.add(strip_vintage(line).lower())
    return keys


def vintage_label(wine):
    """Short vintage label for the breakdown column: '24, '25, or NV."""
    v = wine.get("vintage")
    if v in (None, "", 0):
        return "NV"
    s = str(v)
    return "'" + s[-2:] if s.isdigit() and len(s) >= 2 else s


def _velocity_stats(daily_sales, instock_month_set, on_hand, on_order,
                    window_start_date, window_start_key, current_key, today,
                    lead_time, safety_buffer):
    """
    Compute velocity / runway for one entity (a wine family) from its merged
    daily sales and inventory. Returns a stats dict, or None if it's pure noise
    (no sales, no in-stock evidence, nothing on hand or on order).
    """
    pipeline = on_hand + on_order
    sale_dates = [d for d, q in daily_sales.items()
                  if q > 0 and window_start_date <= d <= today]

    # Active-window start: earliest DATE the wine was selling or, from a
    # snapshot, known in stock. Sale dates strip the leading "not arrived yet"
    # period; snapshot in-stock months pull the start back to catch "arrived but
    # sat unsold" (those slow days are real and must count). created_at is
    # deliberately unused (it's set early at pre-arrival listing).
    start_candidates = list(sale_dates)
    instock_in_window = [mk for mk in instock_month_set
                         if window_start_key <= mk <= current_key]
    if instock_in_window:
        em = min(instock_in_window)
        start_candidates.append(date(em[0], em[1], 1))

    if start_candidates:
        start_date = max(min(start_candidates), window_start_date)
    elif pipeline == 0:
        return None
    else:
        # On order but never seen selling / in stock -- incoming, not reorderable.
        return {
            "total_sales": 0.0, "weekly": 0.0, "monthly": 0.0, "days_in_stock": 0,
            "months_remaining": float("inf"), "reorder": False, "confidence": "",
            "window_start": "-", "window_end": "-",
        }

    # Active-window end: if out of stock now, the wine sold through, so clip to
    # the last day it actually sold. Otherwise the window runs to today.
    if on_hand <= 0 and sale_dates:
        end_date = max(sale_dates)
    else:
        end_date = today
    if end_date < start_date:
        end_date = start_date

    total_sales = sum(q for d, q in daily_sales.items()
                      if start_date <= d <= end_date)
    days_in_stock = (end_date - start_date).days + 1
    days_eff = max(days_in_stock, MIN_VELOCITY_DAYS)
    weekly = total_sales / (days_eff / 7.0)
    monthly = total_sales / (days_eff / DAYS_PER_MONTH)

    months_remaining = (pipeline / monthly) if monthly > 0 else float("inf")
    return {
        "total_sales": total_sales,
        "weekly": weekly,
        "monthly": monthly,
        "days_in_stock": days_in_stock,
        "months_remaining": months_remaining,
        "reorder": months_remaining < (lead_time + safety_buffer),
        "confidence": ("review: thin history"
                       if days_in_stock < THIN_HISTORY_DAYS and total_sales > 0
                       else ""),
        "window_start": start_date.isoformat(),
        "window_end": end_date.isoformat(),
    }


def analyze(wines, inventory, sales, window_months, lead_time, safety_buffer,
            today=None, instock_months=None, discontinued_keys=None):
    """
    Produce a list of per-FAMILY result dicts, sorted most-urgent-first. A family
    is all vintages of the same cuvee; inventory and sales are summed across
    vintages so a low-stock old vintage isn't flagged when its replacement is
    already in stock.
    """
    today = today or date.today()
    instock_months = instock_months or {}
    discontinued_keys = discontinued_keys or set()
    current_key = month_key(today)
    window_start_date = add_months(today.replace(day=1), -(window_months - 1))
    window_start_key = month_key(window_start_date)

    # Index inventory by wine id, summing across warehouses (a wine can sit in
    # more than one warehouse and total sellable stock is the sum).
    inv_by_wine = defaultdict(
        lambda: {"on_hand": 0.0, "on_hold": 0.0, "on_order": 0.0})
    for rec in inventory:
        wid = (rec.get("wine") or {}).get("id")
        if wid is None:
            continue
        inv = rec.get("inventory") or {}
        inv_by_wine[str(wid)]["on_hand"] += to_float(inv.get("on_hand"))
        inv_by_wine[str(wid)]["on_hold"] += to_float(inv.get("on_hold"))
        inv_by_wine[str(wid)]["on_order"] += to_float(inv.get("on_order"))

    # Group active wines into families by their vintage-stripped name.
    families = defaultdict(list)
    for wine in wines:
        if not wine.get("active", True):
            continue
        families[family_name(wine).lower()].append(wine)

    empty_inv = {"on_hand": 0.0, "on_hold": 0.0, "on_order": 0.0}
    results = []
    for fam_key, members in families.items():
        # Merge inventory, daily sales and in-stock months across all vintages.
        on_hand = on_hold = on_order = 0.0
        merged_sales = defaultdict(float)
        merged_instock = set()
        for wine in members:
            wid = str(wine.get("id"))
            inv = inv_by_wine.get(wid, empty_inv)
            on_hand += inv["on_hand"]
            on_hold += inv["on_hold"]
            on_order += inv["on_order"]
            for d, q in sales.get(wid, {}).items():
                merged_sales[d] += q
            merged_instock |= instock_months.get(wid, set())

        # Domestic producers get the short (4-week) lead time.
        producers = {((w.get("producer") or {}).get("name") or "").lower()
                     for w in members}
        fam_lead_time = (DOMESTIC_LEAD_TIME_MONTHS
                         if producers & DOMESTIC_PRODUCERS else lead_time)

        # Discontinued: this wine is on the manual not-reordering list
        # (Vinosmith's discontinued flag isn't exposed in the API).
        discontinued = fam_key in discontinued_keys

        stats = _velocity_stats(
            merged_sales, merged_instock, on_hand, on_order,
            window_start_date, window_start_key, current_key, today,
            fam_lead_time, safety_buffer)
        if stats is None:
            continue

        # Per-vintage on-hand breakdown, most recent vintage first.
        held = [(w, inv_by_wine.get(str(w.get("id")), {}).get("on_hand", 0.0))
                for w in members]
        held = [(w, oh) for w, oh in held if oh > 0]
        held.sort(key=lambda t: (t[0].get("vintage") or 0), reverse=True)
        vintages_on_hand = " · ".join(
            f"{vintage_label(w)}: {round(oh):g}" for w, oh in held)

        # Display name: the family name (vintage stripped), from the newest member.
        newest = max(members, key=lambda w: (w.get("vintage") or 0))
        results.append({
            "confidence": stats["confidence"],
            "name": family_name(newest),
            "weekly_velocity": stats["weekly"],
            "monthly_velocity": stats["monthly"],
            "on_hand": on_hand,
            "on_hold": on_hold,
            "on_order": on_order,
            "vintages_on_hand": vintages_on_hand,
            "months_remaining": stats["months_remaining"],
            "reorder": stats["reorder"] and not discontinued,
            "discontinued": discontinued,
            "weeks_in_stock": round(stats["days_in_stock"] / 7.0, 1),
            "total_sales": stats["total_sales"],
            "window_start": stats["window_start"],
            "window_end": stats["window_end"],
        })

    # Discontinued wines sink to the bottom; above them, most urgent first.
    results.sort(key=lambda r: (r["discontinued"], not r["reorder"],
                                r["months_remaining"]))
    return results


# ---- Excel output -----------------------------------------------------------
COLUMNS = [
    ("Reorder?", "reorder_label", 10),
    ("Wine", "name", 42),
    ("Velocity (btl/wk)", "weekly_velocity", 16),
    ("Velocity (btl/mo)", "monthly_velocity", 16),
    ("On hand", "on_hand", 10),
    ("On hold", "on_hold", 10),
    ("On order", "on_order", 10),
    ("Vintages on hand", "vintages_on_hand", 22),
    ("Months remaining", "months_remaining", 17),
    ("Weeks in stock", "weeks_in_stock", 13),
    ("Confidence", "confidence", 20),
    ("Total sales (window)", "total_sales", 18),
    ("Window start", "window_start", 13),
    ("Window end", "window_end", 12),
]


def write_excel(results, path, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder report"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    reorder_fill = PatternFill("solid", fgColor="F8CBAD")    # warm = needs action
    ok_fill = PatternFill("solid", fgColor="E2EFDA")         # green = fine
    review_fill = PatternFill("solid", fgColor="FFE699")     # yellow = eyeball it
    discontinued_fill = PatternFill("solid", fgColor="D9D9D9")  # grey = discontinued

    # Header on row 1 so the top frozen row is the column headers (no title row
    # for him to delete first).
    header_row = 1
    for ci, (label, _key, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=header_row, column=ci, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    last_row = header_row
    for ri, r in enumerate(results, start=header_row + 1):
        last_row = ri
        row_view = dict(r)
        if r.get("discontinued"):
            row_view["reorder_label"] = "discontinued"
        else:
            row_view["reorder_label"] = "REORDER" if r["reorder"] else "ok"
        mr = r["months_remaining"]
        row_view["months_remaining"] = (
            "∞" if mr == float("inf") else round(mr, 1)
        )
        row_view["weekly_velocity"] = round(r["weekly_velocity"], 1)
        row_view["monthly_velocity"] = round(r["monthly_velocity"], 1)
        row_view["on_hand"] = round(r["on_hand"], 0)
        row_view["on_hold"] = round(r["on_hold"], 0)
        row_view["on_order"] = round(r["on_order"], 0)
        row_view["total_sales"] = round(r["total_sales"], 0)

        if r.get("discontinued"):
            fill = discontinued_fill
        elif r["reorder"]:
            fill = reorder_fill
        else:
            fill = ok_fill
        for ci, (_label, key, _w) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=ri, column=ci, value=row_view.get(key))
            # Highlight the confidence cell when thin so it stands out in-row.
            if key == "confidence" and r.get("confidence"):
                c.fill = review_fill
                c.font = Font(bold=True)
            else:
                c.fill = fill
            if key in ("name", "vintages_on_hand"):
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Run metadata goes at the BOTTOM, two rows below the last wine.
    ws.cell(row=last_row + 2, column=1, value=meta).font = Font(italic=True)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---- Emailing ---------------------------------------------------------------
def email_report(xlsx_path, subject, body):
    """
    Email the report as an attachment via SMTP. All settings come from the
    environment (.env): SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD,
    EMAIL_FROM, EMAIL_TO (comma-separated). Raises if anything is missing.
    """
    import os
    import smtplib
    import ssl
    from email.message import EmailMessage

    host = os.environ.get("SMTP_HOST", "").strip()
    port = int(os.environ.get("SMTP_PORT", "587").strip() or 587)
    user = os.environ.get("SMTP_USER", "").strip()
    password = os.environ.get("SMTP_PASSWORD", "").strip()
    sender = os.environ.get("EMAIL_FROM", user).strip()
    recipients = [a.strip() for a in os.environ.get("EMAIL_TO", "").split(",")
                  if a.strip()]

    missing = [n for n, v in [("SMTP_HOST", host), ("SMTP_USER", user),
                              ("SMTP_PASSWORD", password), ("EMAIL_TO", recipients)]
               if not v]
    if missing:
        raise RuntimeError(
            "Cannot email -- missing in .env: " + ", ".join(missing))

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg.set_content(body)
    with open(xlsx_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=Path(xlsx_path).name)

    context = ssl.create_default_context()
    with smtplib.SMTP(host, port) as server:
        server.starttls(context=context)
        server.login(user, password)
        server.send_message(msg)
    print(f"Emailed {xlsx_path} to {', '.join(recipients)}")


# ---- Entry point ------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--months", type=int, default=DEFAULT_WINDOW_MONTHS,
                    help=f"Trailing window length in months (default {DEFAULT_WINDOW_MONTHS}).")
    ap.add_argument("--lead-time", type=float, default=DEFAULT_LEAD_TIME_MONTHS,
                    help=f"Reorder lead time in months (default {DEFAULT_LEAD_TIME_MONTHS}).")
    ap.add_argument("--safety-buffer", type=float, default=DEFAULT_SAFETY_BUFFER_MONTHS,
                    help=f"Extra safety buffer in months (default {DEFAULT_SAFETY_BUFFER_MONTHS}).")
    ap.add_argument("--output", default=DEFAULT_OUTPUT,
                    help=f"Output .xlsx path (default {DEFAULT_OUTPUT}).")
    ap.add_argument("--end-date", default=None,
                    help="End of the window as YYYY-MM-DD (default: today). "
                         "Use this to include sales dated after the run date.")
    ap.add_argument("--email", action="store_true",
                    help="Email the report to EMAIL_TO using the SMTP_* settings "
                         "in .env (used by the weekly scheduled job).")
    args = ap.parse_args()

    today = parse_date(args.end_date) if args.end_date else date.today()
    if args.end_date and today is None:
        ap.error("--end-date must be YYYY-MM-DD")
    window_start = add_months(today.replace(day=1), -(args.months - 1))

    print(f"Window: {window_start.isoformat()} .. {today.isoformat()} "
          f"({args.months} months)")
    print(f"Lead time {args.lead_time} mo + safety buffer {args.safety_buffer} mo "
          f"= reorder threshold {args.lead_time + args.safety_buffer} mo")

    client = VinosmithClient()
    print("Fetching wines...")
    wines = client.get_wines()
    print(f"  {len(wines)} wines")
    print("Fetching inventory...")
    inventory = client.get_inventory()
    print(f"  {len(inventory)} inventory records")
    print("Fetching supplier orders (chunked)...")
    orders = client.get_supplier_orders(window_start, today)
    print(f"  {len(orders)} supplier orders total")

    # Record today's inventory snapshot, then load the full in-stock history
    # (including the snapshot we just wrote) to refine velocity denominators.
    snap_path = record_snapshot(inventory, today)
    instock_months = load_instock_months()
    print(f"  inventory snapshot recorded -> {snap_path}")

    sales = build_daily_sales(orders)
    discontinued_keys = load_discontinued(DISCONTINUED_PATH)
    print(f"  {len(discontinued_keys)} wines on the discontinued list")
    results = analyze(wines, inventory, sales,
                      window_months=args.months,
                      lead_time=args.lead_time,
                      safety_buffer=args.safety_buffer,
                      today=today,
                      instock_months=instock_months,
                      discontinued_keys=discontinued_keys)

    flagged = sum(1 for r in results if r["reorder"])
    meta = (f"Window {window_start.isoformat()} to {today.isoformat()} "
            f"({args.months} mo)  |  lead {args.lead_time} + buffer "
            f"{args.safety_buffer} = {args.lead_time + args.safety_buffer} mo "
            f"threshold  |  generated {today.isoformat()}")
    write_excel(results, args.output, meta)

    print(f"\n{len(results)} wine families analyzed, {flagged} flagged to reorder.")
    print(f"Wrote {args.output}")

    if args.email:
        subject = (f"TWco reorder report — {today.isoformat()} "
                   f"({flagged} to reorder)")
        body = ("Attached is this week's wine reorder report.\n\n"
                f"{len(results)} wine families analyzed, {flagged} flagged to "
                f"reorder.\n{meta}\n")
        email_report(args.output, subject, body)


if __name__ == "__main__":
    main()
